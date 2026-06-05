"""
mavis.skills.analytics_export — gera CSV / XLSX / PDF a partir de export_rows().

PDF agora aceita `fields_config` para que o usuário escolha:
  - quais KPIs aparecem
  - quais colunas da tabela diária aparecem
  - quais seções extras aparecem (top_destinos, top_equipamentos)

Também expõe `to_pdf_macro()` para o relatório mensal executivo (visão macro).
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional


# ==============================================================
# CONFIG / labels para o seletor de campos do PDF
# ==============================================================
KPI_LABELS: Dict[str, str] = {
    "total_km": "Total KM",
    "total_dias": "Dias úteis",
    "media_km_dia": "Média KM/dia",
    "media_km_semana": "Média KM/semana",
    "litros_estimados": "Litros estimados",
    "custo_combustivel": "Custo combustível (R$)",
    "total_preventivas": "Preventivas",
    "total_atendimentos": "Atendimentos técnicos",
    "total_entregas_insumos": "Entregas de insumos",
    "total_trocas_equipamentos": "Trocas / substituições",
}

COLUMN_LABELS: Dict[str, str] = {
    "data": "Data",
    "km": "KM",
    "visitas": "Visitas",
    "unidades": "Unidades",
    "preventivas": "Prev.",
    "atendimentos": "Atend.",
    "entregas_insumos": "Insumos",
    "trocas": "Trocas",
    "configuracoes": "Config.",
    "equipamentos": "Equipam.",
}

# Larguras das colunas (em mm) para o landscape A4
COLUMN_WIDTHS_MM: Dict[str, float] = {
    "data": 22, "km": 14, "visitas": 16, "unidades": 110,
    "preventivas": 14, "atendimentos": 16, "entregas_insumos": 18,
    "trocas": 16, "configuracoes": 16, "equipamentos": 40,
}

SECTION_LABELS: Dict[str, str] = {
    "kpis": "Indicadores",
    "top_destinos": "Top destinos",
    "top_equipamentos": "Top equipamentos",
    "diario": "Diário",
}

DEFAULT_PDF_CONFIG: Dict[str, List[str]] = {
    "kpis": list(KPI_LABELS.keys()),
    "columns": ["data", "km", "visitas", "unidades", "preventivas",
                "atendimentos", "entregas_insumos", "trocas"],
    "sections": ["kpis", "top_destinos", "diario"],
}


def pdf_fields_catalog() -> Dict[str, Any]:
    """Retorna o catálogo completo para o frontend montar o modal de seleção."""
    return {
        "kpis": [{"key": k, "label": v} for k, v in KPI_LABELS.items()],
        "columns": [{"key": k, "label": v} for k, v in COLUMN_LABELS.items()],
        "sections": [{"key": k, "label": v} for k, v in SECTION_LABELS.items()],
        "defaults": DEFAULT_PDF_CONFIG,
    }


def _resolve_fields(fields: Optional[Dict[str, Any]]) -> Dict[str, List[str]]:
    """Mescla a config recebida com os defaults, validando chaves."""
    base = {k: list(v) for k, v in DEFAULT_PDF_CONFIG.items()}
    if not fields:
        return base
    if isinstance(fields.get("kpis"), list):
        base["kpis"] = [k for k in fields["kpis"] if k in KPI_LABELS]
    if isinstance(fields.get("columns"), list):
        base["columns"] = [c for c in fields["columns"] if c in COLUMN_LABELS]
        if not base["columns"]:
            base["columns"] = DEFAULT_PDF_CONFIG["columns"][:]
    if isinstance(fields.get("sections"), list):
        base["sections"] = [s for s in fields["sections"] if s in SECTION_LABELS]
    return base


# ==============================================================
# CSV
# ==============================================================
def to_csv(rows: List[Dict[str, Any]]) -> bytes:
    if not rows:
        return "data,km,visitas,unidades\n".encode("utf-8-sig")
    fieldnames = list(rows[0].keys())
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=";")
    writer.writeheader()
    for r in rows:
        writer.writerow(r)
    return buf.getvalue().encode("utf-8-sig")


# ==============================================================
# XLSX
# ==============================================================
def to_xlsx(rows: List[Dict[str, Any]], kpis: Optional[Dict[str, Any]] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Diário"

    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(bold=True, color="0A0A0A", name="Calibri", size=11)
    border = Border(*(Side(style="thin", color="DDDDDD"),) * 4)

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for col_idx, h in enumerate(headers, 1):
            max_len = max([len(str(h))] + [len(str(r.get(h, ""))) for r in rows])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    else:
        ws.append(["Sem dados no filtro selecionado."])

    if kpis:
        ws2 = wb.create_sheet("KPIs")
        ws2.append(["Indicador", "Valor"])
        ws2["A1"].font = header_font
        ws2["A1"].fill = header_fill
        ws2["B1"].font = header_font
        ws2["B1"].fill = header_fill
        skip = {"top_destinos", "top_equipamentos", "filtro"}
        for k, v in kpis.items():
            if k in skip:
                continue
            ws2.append([k, v])
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 24

        if kpis.get("top_destinos"):
            ws3 = wb.create_sheet("Top Destinos")
            ws3.append(["Unidade", "Visitas"])
            ws3["A1"].font = header_font
            ws3["A1"].fill = header_fill
            ws3["B1"].font = header_font
            ws3["B1"].fill = header_fill
            for d in kpis["top_destinos"]:
                ws3.append([d["unidade"], d["visitas"]])
            ws3.column_dimensions["A"].width = 48
            ws3.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==============================================================
# PDF (com seleção de campos)
# ==============================================================
def to_pdf(rows: List[Dict[str, Any]], kpis: Optional[Dict[str, Any]] = None,
           filtro: Optional[Dict[str, Any]] = None,
           fields: Optional[Dict[str, Any]] = None) -> bytes:
    """Gera o PDF respeitando a seleção de campos do usuário.

    `fields` (opcional) = {"kpis":[...], "columns":[...], "sections":[...]}.
    Se None, usa DEFAULT_PDF_CONFIG (comportamento legado: PDF completo).
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    cfg = _resolve_fields(fields)
    sel_kpis = cfg["kpis"]
    sel_cols = cfg["columns"]
    sel_secs = set(cfg["sections"])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title="Mavis · Relatório Analytics")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=colors.HexColor("#F59E0B"),
                                 alignment=0, fontSize=18, spaceAfter=6)
    subtitle = ParagraphStyle("sub", parent=styles["Normal"],
                              textColor=colors.HexColor("#71717a"),
                              fontSize=9, spaceAfter=10)
    section = ParagraphStyle("sec", parent=styles["Heading2"],
                             textColor=colors.HexColor("#F59E0B"),
                             fontSize=11, spaceAfter=6, spaceBefore=10)

    story = []
    story.append(Paragraph("Mavis · Relatório Analytics", title_style))
    fstr_parts = []
    if filtro:
        if filtro.get("start"):
            fstr_parts.append(f"De {filtro['start']}")
        if filtro.get("end"):
            fstr_parts.append(f"até {filtro['end']}")
        if filtro.get("unidade"):
            fstr_parts.append(f"unidade: {filtro['unidade']}")
    fstr_parts.append(f"gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    story.append(Paragraph(" · ".join(fstr_parts), subtitle))

    # ---------- Seção KPIs ----------
    if kpis and "kpis" in sel_secs and sel_kpis:
        story.append(Paragraph(SECTION_LABELS["kpis"], section))
        kpi_pairs: List[Tuple[str, Any]] = [(KPI_LABELS[k], kpis.get(k, 0)) for k in sel_kpis
                                            if k in KPI_LABELS]
        # Em grid de 5 colunas
        cols = 5
        rows_kpi = []
        for i in range(0, len(kpi_pairs), cols):
            chunk = kpi_pairs[i:i + cols]
            # Padding para fechar a linha
            while len(chunk) < cols:
                chunk.append(("", ""))
            labels = [c[0] for c in chunk]
            vals = [str(c[1]) for c in chunk]
            rows_kpi.append(labels)
            rows_kpi.append(vals)
        if rows_kpi:
            t = Table(rows_kpi, colWidths=[52 * mm] * cols)
            style_cmds = [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#27272a")),
                ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#EEEEEE")),
            ]
            # Backgrounds nas linhas de labels (linhas pares: 0, 2, 4...)
            for li in range(0, len(rows_kpi), 2):
                style_cmds.append(("BACKGROUND", (0, li), (-1, li), colors.HexColor("#FEF3C7")))
            # Bold + amber nas linhas de valores (ímpares: 1, 3, 5...)
            for li in range(1, len(rows_kpi), 2):
                style_cmds.append(("FONTNAME", (0, li), (-1, li), "Helvetica-Bold"))
                style_cmds.append(("TEXTCOLOR", (0, li), (-1, li), colors.HexColor("#F59E0B")))
            t.setStyle(TableStyle(style_cmds))
            story.append(t)
            story.append(Spacer(1, 4 * mm))

    # ---------- Top destinos ----------
    if kpis and "top_destinos" in sel_secs:
        td = kpis.get("top_destinos") or []
        if td:
            story.append(Paragraph(SECTION_LABELS["top_destinos"], section))
            data = [["#", "Unidade", "Visitas"]]
            for i, d in enumerate(td, 1):
                data.append([str(i), d["unidade"], str(d["visitas"])])
            tbl = Table(data, colWidths=[10 * mm, 160 * mm, 25 * mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEEEEE")),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 4 * mm))

    # ---------- Top equipamentos ----------
    if kpis and "top_equipamentos" in sel_secs:
        te = kpis.get("top_equipamentos") or []
        if te:
            story.append(Paragraph(SECTION_LABELS["top_equipamentos"], section))
            data = [["#", "Item", "Qtd"]]
            for i, d in enumerate(te, 1):
                data.append([str(i), d["item"], str(d["qtd"])])
            tbl = Table(data, colWidths=[10 * mm, 160 * mm, 25 * mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEEEEE")),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 4 * mm))

    # ---------- Diário ----------
    if rows and "diario" in sel_secs and sel_cols:
        story.append(Paragraph(SECTION_LABELS["diario"], section))
        headers = [COLUMN_LABELS[c] for c in sel_cols]
        data = [headers]
        for r in rows:
            row_vals = []
            for c in sel_cols:
                v = r.get(c, "")
                if c == "unidades":
                    v = (v or "")[:80]
                row_vals.append(str(v) if v != "" else "")
            data.append(row_vals)
        col_widths = [COLUMN_WIDTHS_MM.get(c, 20) * mm for c in sel_cols]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        # Alinhamento à direita para colunas numéricas
        numeric_cols = {"km", "visitas", "preventivas", "atendimentos",
                        "entregas_insumos", "trocas", "configuracoes"}
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEEEEE")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]
        for col_idx, c in enumerate(sel_cols):
            if c in numeric_cols:
                style_cmds.append(("ALIGN", (col_idx, 0), (col_idx, -1), "RIGHT"))
        tbl.setStyle(TableStyle(style_cmds))
        story.append(tbl)
    elif "diario" in sel_secs and not rows:
        story.append(Paragraph("Sem registros para o filtro selecionado.", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


# ==============================================================
# PDF MACRO MENSAL (visão executiva, sem foco em KM)
# ==============================================================
def to_pdf_macro(macro: Dict[str, Any], narrativa: str = "") -> bytes:
    """PDF do resumo mensal MACRO. Foco em volume e tipo de operação, não em KM."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm,
                            title=f"Mavis · Resumo Mensal {macro.get('month','')}")
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"],
                           textColor=colors.HexColor("#F59E0B"),
                           fontSize=20, alignment=0, spaceAfter=4)
    subtitle = ParagraphStyle("sub", parent=styles["Normal"],
                              textColor=colors.HexColor("#71717a"),
                              fontSize=9, spaceAfter=10)
    section = ParagraphStyle("sec", parent=styles["Heading2"],
                             textColor=colors.HexColor("#F59E0B"),
                             fontSize=12, spaceAfter=6, spaceBefore=10)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10,
                          leading=14, spaceAfter=4)

    story = []
    story.append(Paragraph(f"Resumo Mensal · {macro.get('month', '')}", title))
    story.append(Paragraph(
        f"Visão macro operacional · comparativo vs {macro.get('previous_month', '—')} · "
        f"gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle))

    cur = macro.get("current") or {}
    prev = macro.get("previous") or {}
    d = macro.get("deltas") or {}

    def fmt_delta(v: Optional[float]) -> str:
        if v is None:
            return "—"
        sign = "+" if v > 0 else ""
        return f"{sign}{v}%"

    # Tabela comparativa operacional (NÃO inclui KM nem combustível)
    story.append(Paragraph("Operação no mês", section))
    table_data = [
        ["Indicador", "Mês atual", "Mês anterior", "Variação"],
        ["Preventivas", cur.get("preventivas", 0), prev.get("preventivas", 0), fmt_delta(d.get("preventivas"))],
        ["Atendimentos técnicos", cur.get("atendimentos", 0), prev.get("atendimentos", 0), fmt_delta(d.get("atendimentos"))],
        ["Entregas de insumos", cur.get("entregas", 0), prev.get("entregas", 0), fmt_delta(d.get("entregas"))],
        ["Trocas / substituições", cur.get("trocas", 0), prev.get("trocas", 0), fmt_delta(d.get("trocas"))],
        ["Configurações", cur.get("configuracoes", 0), prev.get("configuracoes", 0), fmt_delta(d.get("configuracoes"))],
        ["Visitas totais", cur.get("visitas", 0), prev.get("visitas", 0), fmt_delta(d.get("visitas"))],
        ["Dias úteis trabalhados", cur.get("dias_uteis", 0), prev.get("dias_uteis", 0), fmt_delta(d.get("dias_uteis"))],
    ]
    tbl = Table([[str(x) for x in row] for row in table_data],
                colWidths=[80 * mm, 30 * mm, 35 * mm, 30 * mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEEEEE")),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 5 * mm))

    # Top 5 unidades
    top_un = cur.get("top_unidades") or []
    if top_un:
        story.append(Paragraph("Top 5 unidades visitadas", section))
        ud = [["#", "Unidade", "Visitas"]]
        for i, u in enumerate(top_un, 1):
            ud.append([str(i), u["unidade"], str(u["visitas"])])
        t = Table(ud, colWidths=[10 * mm, 135 * mm, 30 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEEEEE")),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ]))
        story.append(t)
        story.append(Spacer(1, 4 * mm))

    # Top 3 equipamentos
    top_eq = cur.get("top_equipamentos") or []
    if top_eq:
        story.append(Paragraph("Top 3 equipamentos manuseados", section))
        ed = [["#", "Item", "Quantidade"]]
        for i, e in enumerate(top_eq, 1):
            ed.append([str(i), e["item"], str(e["qtd"])])
        t = Table(ed, colWidths=[10 * mm, 135 * mm, 30 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.2, colors.HexColor("#EEEEEE")),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ]))
        story.append(t)
        story.append(Spacer(1, 4 * mm))

    # Concentração no top3 (insight macro)
    conc = macro.get("concentracao_top3", 0)
    if conc:
        story.append(Paragraph("Concentração operacional", section))
        story.append(Paragraph(
            f"As <b>3 unidades mais visitadas</b> concentraram "
            f"<b>{round(conc * 100, 1)}%</b> do total de visitas no mês.", body))
        story.append(Spacer(1, 3 * mm))

    # Narrativa Gemini
    if narrativa:
        story.append(Paragraph("Análise executiva", section))
        for para in narrativa.split("\n\n"):
            if para.strip():
                # Escapa < e > exceto tags pretendidas
                safe = (para.replace("&", "&amp;")
                            .replace("<", "&lt;").replace(">", "&gt;")
                            .replace("\n", "<br/>"))
                story.append(Paragraph(safe, body))

    doc.build(story)
    return buf.getvalue()
